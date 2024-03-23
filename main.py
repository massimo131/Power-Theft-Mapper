# Importing necessary modules
import openpyxl
from openpyxl.reader.excel import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import gmplot
from geopy.distance import geodesic
import math

# Loading the Excel workbook
wrkbk = load_workbook("test.xlsx")
sh = wrkbk.active

# Extracting cell value for distance calculation
c = sh.cell(row=3, column=1)
lth = float(c.value)

# Initial latitude and longitude values
lat3 = 19.198163880068083
lat4 = float(lat3)
lon3 = 72.82592142624505
lon4 = float(lon3)

# Initial bearing value
bear = 141.48796990531704
bearing = float(bear)

# Earth radius
rr = 6371
R = float(rr)

# Converting bearing to radians
brng = math.radians(bearing)

# Converting latitude and longitude to radians
lat1 = math.radians(lat4)
lon1 = math.radians(lon4)

# Calculating new latitude and longitude based on distance and bearing
lat2 = math.asin(math.sin(lat1) * math.cos(lth / R) + math.cos(lat1) * math.sin(lth / R) * math.cos(brng))
lon2 = lon1 + math.atan2(math.sin(brng) * math.sin(lth / R) * math.cos(lat1), math.cos(lth / R) - math.sin(lat1) * math.sin(lat2))

# Converting radians to degrees
lat6 = math.degrees(lat2)
lon6 = math.degrees(lon2)

# Coordinates for plotting on the map
lat = [19.198163880068083, lat6, 19.117338429973362]
long = [72.82592142624505, lon6, 72.8935775820629]

# Creating a Google Map plotter object
gmapOne = gmplot.GoogleMapPlotter(19.198163880068083, 72.82592142624505, 15)

# Scatter plotting the coordinates
gmapOne.scatter(lat, long, size=50, marker=False)

# Plotting the route
gmapOne.plot(lat, long, 'red', edge_width=2.5)

# Saving the map to an HTML file
gmapOne.draw("map.html")

# Email Configuration
fromaddr = "sender@abc.com"
toaddr = "receiver@abc.com"

# Creating the email message
msg = MIMEMultipart()
msg['From'] = fromaddr
msg['To'] = toaddr
msg['Subject'] = "Location of Power Theft"

# Email body
body = "Visit this location and resolve the problem of power theft."
msg.attach(MIMEText(body, 'plain'))

# Attaching the map HTML file to the email
filename = "map.html"
attachment = open("map.html", "rb")
p = MIMEBase('application', 'octet-stream')
p.set_payload((attachment).read())
encoders.encode_base64(p)
p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
msg.attach(p)

# SMTP Configuration and sending the email
s = smtplib.SMTP('smtp.gmail.com', 587)
s.starttls()
s.login(fromaddr, "password")
text = msg.as_string()
s.sendmail(fromaddr, toaddr, text)
s.quit()